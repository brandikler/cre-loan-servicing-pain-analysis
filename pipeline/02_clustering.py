"""
Stage 3 — Semantic clustering
==============================

This script is a RECORD of the actual pipeline run that produced
clusters.json and pains_with_clusters.xlsx, NOT a re-runnable
API-calling pipeline.

The clustering was performed via direct LLM reasoning at n=47, NOT via
embedding-based clustering (HDBSCAN, KMeans). At this corpus size, manual
LLM clustering produces tighter, more interpretable themes than
ML-based methods, which over-split or generate noisy clusters. The
trade-off: less reproducible. A different LLM run might split or merge
clusters slightly differently. The thematic finding is robust because
corroboration is at the source level, not the cluster level.

Cluster definitions (the CLUSTERS dict below) were assigned by reading
each pain in extraction_results.json and grouping by semantic similarity.
The 13 clusters were then scored by:

  composite_score = severity_weighted_sum + (unique_source_ids x unique_source_types)

where severity weights are: high=3, medium=2, low=1.

To re-execute the cluster ASSIGNMENT against an updated extraction:
  1. Read extraction_results.json
  2. For each pain, prompt an LLM to assign a cluster ID, providing
     the existing cluster definitions and labels.
  3. (Recommended: add a separate embedding-based clustering pass as a
     check on manual assignments — see README "What I'd do differently".)

Outputs:
  - clusters.json
  - ranked_themes.md
  - pains_with_clusters.xlsx
"""

"""
Cluster the 47 extracted pains into themes, score them, and produce ranked output.
Method: manual semantic clustering by LLM (transparent — at n=47 this beats ML clustering).
"""
import json
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

with open('/sessions/kind-hopeful-edison/mnt/outputs/extraction_results.json') as f:
    results = json.load(f)

# Build the flat pain list with sequential pain_ids matching the flat XLSX
pains = []
pid = 0
for r in results:
    for p in r['pains']:
        pid += 1
        pains.append({
            'pain_id': f'P{pid:03d}',
            'source_id': r['source_id'],
            'source_row': r['row'],
            'source_type': r['source_type'],
            'persona_fit': r['persona_fit'],
            **p
        })

# ============================================================
# Cluster definitions (manual semantic assignment)
# ============================================================
CLUSTERS = {
    "C1": {
        "label": "Manual / bespoke reporting & analytics",
        "description": "Reports developed per request rather than pulled from system; no self-serve reporting; KPI/portfolio reports require manual compilation, eng involvement, or SQL workarounds.",
        "members": ["P007", "P009", "P013", "P018", "P023", "P027", "P038"]
    },
    "C2": {
        "label": "Reconciliation drift between systems & reports",
        "description": "Numbers don't match between internal books, capital provider reports, and operational systems. Reconciliation procedures are home-grown; trust in numbers erodes; this drives system replacement.",
        "members": ["P024", "P025", "P045", "P047"]
    },
    "C3": {
        "label": "Manual covenant / event / trigger monitoring",
        "description": "Covenant compliance, watchlist signals, escrow events, and risk triggers are monitored manually rather than via automated alerts. Compliance and risk-surfacing scale with headcount.",
        "members": ["P010", "P015", "P019"]
    },
    "C4": {
        "label": "Manual data ingest from borrowers, contracts & counterparties",
        "description": "No self-service borrower portal; property/financial updates collected via correspondence; contract data manually re-keyed; data tapes manually generated. Inbound data is unstructured and must be normalized by analysts.",
        "members": ["P006", "P008", "P017", "P022"]
    },
    "C5": {
        "label": "Loan boarding handoff broken",
        "description": "Origination → servicing data flow is broken or absent; loan boarding requires manual handoff for both originated and acquired loans.",
        "members": ["P014"]
    },
    "C6": {
        "label": "Vendor support quality, responsiveness & release management",
        "description": "Vendor support is slow, unhelpful, or argumentative; bug fixes take 6+ months; release management is poor (no rollback, undisclosed server moves, customers run UAT for vendor releases). Resolution delays bleed over reporting periods.",
        "members": ["P003", "P020", "P029", "P033", "P036", "P037", "P043"]
    },
    "C7": {
        "label": "UI/UX complexity & long learning curve",
        "description": "Interfaces are ancient, unintuitive, or weirdly structured. Even experienced operations personnel struggle with simple tasks. Long learning curves drive workflow friction.",
        "members": ["P004", "P031", "P032", "P034"]
    },
    "C8": {
        "label": "Customization & extensibility barriers (APIs, modeling, fit)",
        "description": "Cannot customize platform to fit needs; APIs poorly documented; programming languages opaque; even simple changes require developer support; platforms don't fit commercial workflows; no native financial modeling forces Excel as middleware.",
        "members": ["P001", "P002", "P011", "P041", "P042", "P044"]
    },
    "C9": {
        "label": "Manual money-event workflows (draws, disbursements, restructures)",
        "description": "Construction draws, disbursements, adjustments, restructures, and recoveries require manual coordination across multiple external parties and don't have unified workflow tooling.",
        "members": ["P012", "P046"]
    },
    "C10": {
        "label": "Manual compliance auditing",
        "description": "Compliance auditing is manual file review with no automated rules embedded in the platform; compliance reporting across capital providers is fragmented and duplicative.",
        "members": ["P021", "P026"]
    },
    "C11": {
        "label": "Code quality, performance & hosting",
        "description": "Code held together by 'duct tape' with null pointer exceptions; app-server latency unresolved without escalation; shared hosting environments where one customer affects another.",
        "members": ["P028", "P035", "P040"]
    },
    "C12": {
        "label": "Pricing, missing features & procurement gaps",
        "description": "Concerns around price-to-value; missing modules (GL); no trial period or pre-contract documentation. Customers paying for unused features.",
        "members": ["P005", "P030", "P039"]
    },
    "C13": {
        "label": "Process gap acknowledgment (meta)",
        "description": "Organizations explicitly staff against 'process improvement' — not a workflow-specific pain but a meta-signal that current workflows have material gaps.",
        "members": ["P016"]
    },
}

# Build pain_id → cluster lookup
pain_to_cluster = {}
for cid, cdata in CLUSTERS.items():
    for pid in cdata['members']:
        pain_to_cluster[pid] = cid

# Sanity check
expected = {p['pain_id'] for p in pains}
assigned = set(pain_to_cluster.keys())
missing = expected - assigned
extra = assigned - expected
assert not missing, f"Unassigned pains: {missing}"
assert not extra, f"Extra pains in clusters: {extra}"

# ============================================================
# Score each cluster
# ============================================================
SEVERITY_WEIGHT = {'high': 3, 'medium': 2, 'low': 1, None: 1}

cluster_scores = {}
for cid, cdata in CLUSTERS.items():
    member_pains = [p for p in pains if p['pain_id'] in cdata['members']]
    sev_sum = sum(SEVERITY_WEIGHT[p['severity_signal']] for p in member_pains)
    unique_sources = len(set(p['source_id'] for p in member_pains))
    unique_source_types = len(set(p['source_type'] for p in member_pains))
    high_count = sum(1 for p in member_pains if p['severity_signal'] == 'high')

    cluster_scores[cid] = {
        'label': cdata['label'],
        'description': cdata['description'],
        'pain_count': len(member_pains),
        'severity_weighted': sev_sum,
        'high_severity_count': high_count,
        'unique_sources': unique_sources,
        'unique_source_types': unique_source_types,
        'cross_source_score': unique_sources * unique_source_types,
        'composite_score': sev_sum + (unique_sources * unique_source_types),
        'members': cdata['members'],
        'pain_summaries': [p['pain_summary'] for p in member_pains],
    }

# Rank by composite score (severity + cross-source corroboration)
ranked = sorted(cluster_scores.items(), key=lambda x: x[1]['composite_score'], reverse=True)

# ============================================================
# OUTPUT 1: clusters.json
# ============================================================
with open('/sessions/kind-hopeful-edison/mnt/outputs/clusters.json', 'w') as f:
    json.dump({
        'methodology': {
            'method': 'Manual semantic clustering by LLM',
            'rationale': 'At n=47, LLM-driven clustering produces tighter, more interpretable themes than ML-based approaches (HDBSCAN/KMeans on embeddings). Trade-off: less reproducible without an explicit cluster prompt.',
            'scoring': 'composite_score = severity_weighted_sum + (unique_source_ids × unique_source_types). Severity weights: high=3, medium=2, low=1.',
            'limitations': 'JD-derived pains over-index in severity (workflow descriptions read as blocking). Nortridge over-indexed in raw count (8 of 40 source rows). Discount Nortridge-only clusters when interpreting.'
        },
        'rankings': [
            {'rank': i+1, 'cluster_id': cid, **scores}
            for i, (cid, scores) in enumerate(ranked)
        ]
    }, f, indent=2)

# ============================================================
# OUTPUT 2: Markdown ranked themes report
# ============================================================
md = ['# Ranked Themes — CRE Loan Servicing Analyst Pain Synthesis\n']
md.append(f'**Corpus:** 40 source rows → 47 extracted pains → {len(CLUSTERS)} clusters\n')
md.append('**Scoring:** `composite_score = severity_weighted + (unique_source_ids × unique_source_types)`\n')
md.append('Severity weights: high=3, medium=2, low=1.\n')
md.append('---\n')

for i, (cid, s) in enumerate(ranked, 1):
    member_pains = [p for p in pains if p['pain_id'] in s['members']]
    src_breakdown = Counter(p['source_id'] for p in member_pains)
    type_breakdown = Counter(p['source_type'] for p in member_pains)
    md.append(f'## #{i} — {s["label"]} (`{cid}`)\n')
    md.append(f'**Composite score:** {s["composite_score"]} | **Pains:** {s["pain_count"]} | **Severity-weighted:** {s["severity_weighted"]} | **High-severity:** {s["high_severity_count"]}\n')
    md.append(f'**Unique sources:** {s["unique_sources"]} | **Unique source types:** {s["unique_source_types"]} (cross-source score: {s["cross_source_score"]})\n')
    md.append(f'\n_{s["description"]}_\n')
    md.append(f'\n**Source breakdown:** {dict(src_breakdown)}')
    md.append(f'**Type breakdown:** {dict(type_breakdown)}\n')
    md.append('**Member pains:**')
    for p in member_pains:
        md.append(f'- `{p["pain_id"]}` [{p["severity_signal"]}] **{p["source_id"]}** (row {p["source_row"]}, {p["source_type"]}): {p["pain_summary"]}')
    md.append('\n---\n')

with open('/sessions/kind-hopeful-edison/mnt/outputs/ranked_themes.md', 'w') as f:
    f.write('\n'.join(md))

# ============================================================
# OUTPUT 3: Updated flat XLSX with cluster column
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pains_with_Clusters"

headers = [
    'pain_id', 'cluster_id', 'cluster_label', 'rank',
    'source_id', 'source_row', 'source_type', 'persona_fit',
    'pain_summary', 'severity_signal', 'context_tag',
    'user_segment', 'dollar_signal', 'proposed_solution', 'verbatim_quotes'
]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    cell.alignment = Alignment(vertical='center', wrap_text=True)

# rank lookup
rank_lookup = {cid: i+1 for i, (cid, _) in enumerate(ranked)}

for p in pains:
    cid = pain_to_cluster[p['pain_id']]
    ws.append([
        p['pain_id'],
        cid,
        CLUSTERS[cid]['label'],
        rank_lookup[cid],
        p['source_id'],
        p['source_row'],
        p['source_type'],
        p['persona_fit'],
        p['pain_summary'],
        p['severity_signal'],
        p['context_tag'],
        p['user_segment'],
        p['dollar_signal'],
        p['proposed_solution'],
        ' || '.join(p['verbatim_quotes'])
    ])

widths = [8, 8, 40, 6, 28, 8, 16, 8, 50, 10, 16, 28, 28, 28, 80]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Add Rankings sheet
ws2 = wb.create_sheet("Rankings")
ws2.append([
    'rank', 'cluster_id', 'cluster_label', 'composite_score',
    'pain_count', 'severity_weighted', 'high_severity_count',
    'unique_sources', 'unique_source_types', 'description'
])
for col_idx in range(1, 11):
    cell = ws2.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    cell.alignment = Alignment(vertical='center', wrap_text=True)

for i, (cid, s) in enumerate(ranked, 1):
    ws2.append([
        i, cid, s['label'], s['composite_score'],
        s['pain_count'], s['severity_weighted'], s['high_severity_count'],
        s['unique_sources'], s['unique_source_types'], s['description']
    ])

widths2 = [6, 8, 40, 12, 8, 12, 8, 8, 8, 80]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = 'A2'

wb.save('/sessions/kind-hopeful-edison/mnt/outputs/pains_with_clusters.xlsx')

print('=== Ranked themes ===')
for i, (cid, s) in enumerate(ranked, 1):
    print(f'{i}. {cid} {s["label"]}')
    print(f'   composite={s["composite_score"]} pains={s["pain_count"]} sev_wt={s["severity_weighted"]} high={s["high_severity_count"]} sources={s["unique_sources"]} src_types={s["unique_source_types"]}')
print()
print(f'Total clusters: {len(CLUSTERS)}')
print(f'Total pains assigned: {sum(len(c["members"]) for c in CLUSTERS.values())}')
