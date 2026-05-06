"""
Stage 2 — Per-document extraction
=================================

This script is a RECORD of the actual pipeline run that produced
extraction_results.json, NOT a re-runnable API-calling pipeline.

The Stage 2 extraction was executed via direct LLM reasoning (Claude
Sonnet 4.5) using the prompt in /prompts/extraction_prompt.md applied
to each row of /data/corpus.xlsx. The structured outputs (47 pains
across 38 of 40 sources) are encoded as Python data structures in
the `results` list below.

To re-execute against an updated corpus:
  1. Read /data/corpus.xlsx
  2. For each row:
       - if source_type == 'job_description': source_text = row['quick_note']
       - else: source_text = row['raw_text']
  3. Send to Claude API with the extraction prompt; parse the JSON output.
  4. Aggregate into results structure matching extraction_results.json.

The two-pipeline routing (job_description -> quick_note, else raw_text)
is non-negotiable. Running the prompt against raw JD text returns zero
pains because JDs describe duties neutrally rather than expressing
friction. See /eval/eval_results.md for the discovery and fix.

Outputs:
  - extraction_results.json  : per-source nested structure
  - extraction_pains_flat.xlsx : one row per pain, review-friendly
"""

"""
Builds the structured pain extraction dataset from the corpus.
Output: extraction_results.json (per-source) + extraction_pains_flat.xlsx (per-pain).
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# EXTRACTION RESULTS — manual application of the extraction prompt
# Schema per source: {source_id, row, source_type, pains: [...]}
# Schema per pain: {pain_summary, verbatim_quotes, severity_signal,
#                   dollar_signal, user_segment, proposed_solution, context_tag}
# ============================================================

results = [
    # =========================================================
    # ABRIGO/SAGEWORKS - ForumAggregate
    # =========================================================
    {
        "source_id": "Abrigo/Sageworks",
        "row": 2, "source_type": "ForumAggregate", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Tedious for commercial; geared toward consumer lending",
                "verbatim_quotes": [
                    "Agree on the tedious comment as it seems geared much more towards consumer lending"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Commercial lending user",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },

    # =========================================================
    # FINASTRA LOAN IQ - Reviews
    # =========================================================
    {
        "source_id": "Finastra LoanIQ", "row": 3, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Cannot fully customize product to needs",
                "verbatim_quotes": [
                    "I cannot customize fully to how I want the product",
                    "Based on all the teams that use this product, we would've liked the product to have more customization, however that is not the case"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Multi-team user organization",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Finastra LoanIQ", "row": 4, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Vendor issue resolution slow; delays bleed over reporting periods",
                "verbatim_quotes": [
                    "we have to go back to Finastra to fix issues and the resolutions are not as quick as we need",
                    "This often bleeds over reporting periods which is problematic"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Daily user at very large enterprise consolidating loan systems; works on non-traditional loans with customized features",
                "proposed_solution": None,
                "context_tag": "support"
            },
            {
                "pain_summary": "Complex UI; experienced ops personnel struggle with simple tasks",
                "verbatim_quotes": [
                    "I find the research to be difficult and not intuitive",
                    "It took me a long time to figure out where to find things",
                    "I know our Operational personnel struggle to learn all of it and there is a long learning curve with it",
                    "We have some experienced ops personnel that do not know how to fairly simple tasks due to the complicated interface"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Operational personnel at large enterprise",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Finastra LoanIQ", "row": 5, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Pricing concern (hedged)",
                "verbatim_quotes": [
                    "If you don't consider price perspective, then loan IQ is a good option to use"
                ],
                "severity_signal": "low",
                "dollar_signal": "If you don't consider price perspective",
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "pricing"
            }
        ]
    },

    # =========================================================
    # JD - CRE ANALYST (QAR) - 6 rows, Quick Note as source
    # =========================================================
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 6, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "No self-service borrower portal; analyst processes requests manually",
                "verbatim_quotes": [
                    "Borrower requests and consents are manually processed by the analyst — no self-service borrower portal",
                    "Borrower interactions are inbound work for the servicing analyst"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 7, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Quarterly Asset Reports manually compiled; significant analyst hours per quarter",
                "verbatim_quotes": [
                    "QARs are manually compiled each quarter — significant analyst hours consumed in producing investor/lender reporting",
                    "Implies no real-time investor portal or automated QAR generation"
                ],
                "severity_signal": "high",
                "dollar_signal": "significant analyst hours consumed",
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 8, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Property data collected via correspondence; unstructured inbound data",
                "verbatim_quotes": [
                    "Property-level updates collected via correspondence (email/phone) rather than borrower data feed or portal",
                    "Inbound borrower data is unstructured and must be normalized by the analyst"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "integration"
            }
        ]
    },
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 9, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "No real-time portfolio dashboard; weekly updates manually compiled",
                "verbatim_quotes": [
                    "Weekly portfolio updates are manually compiled — no real-time portfolio dashboard surfacing issues, discussions, and events automatically",
                    "Analyst is the aggregation layer"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 10, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Covenant/event/trigger monitoring is manual; no automated alerts",
                "verbatim_quotes": [
                    "Event/trigger monitoring (covenant breaches, watchlist signals, escrow shortfalls, payment events) is a human job rather than automated alerts",
                    "Covenant/trigger automation is missing or partial"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - CRE Analyst (QAR)", "row": 11, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Servicing platforms can't model financials; Excel is de facto modeling layer",
                "verbatim_quotes": [
                    "CRE financial analysis lives in Excel models *alongside* the servicing platform",
                    "servicing platforms don't natively support analyst-level financial modeling (DSCR, NPV, IRR scenarios on stressed cash flows)",
                    "Excel is the de facto modeling layer"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "CRE servicing analyst",
                "proposed_solution": None,
                "context_tag": "integration"
            }
        ]
    },

    # =========================================================
    # JD - PEACHTREE - 5 rows
    # =========================================================
    {
        "source_id": "JD - Peachtree Loan Servicing Analyst", "row": 12, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Draw processing requires manual coordination across 4+ external parties",
                "verbatim_quotes": [
                    "Draw processing requires manual coordination across 4+ external parties (borrowers, construction monitors, participants, lenders)",
                    "No unified workflow tool — coordination happens via email/calls/spreadsheets"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing analyst handling construction draws",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Peachtree Loan Servicing Analyst", "row": 13, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Reports developed per request, not pulled from system",
                "verbatim_quotes": [
                    "Reports are *developed* per request, not pulled from system",
                    "Reporting layer of the servicing platform is too weak for analyst needs; bespoke report-building is part of the job"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Peachtree Loan Servicing Analyst", "row": 14, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Loan boarding manual handoff; origination → servicing data flow broken",
                "verbatim_quotes": [
                    "Loan boarding is a manual handoff",
                    "Origination → servicing data flow is broken or absent; analyst manually facilitates onboarding for both originated and acquired loans"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing analyst",
                "proposed_solution": None,
                "context_tag": "onboarding"
            }
        ]
    },
    {
        "source_id": "JD - Peachtree Loan Servicing Analyst", "row": 15, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Covenant compliance is ad-hoc manual review; no automated tracking",
                "verbatim_quotes": [
                    "Covenant compliance is \"ad-hoc review\" — not automated covenant tracking",
                    "Analyst reads loan docs to verify compliance manually",
                    "Direct covenant-tracking pain"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Peachtree Loan Servicing Analyst", "row": 16, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Process improvement is staffed against; current processes have material gaps",
                "verbatim_quotes": [
                    "Process improvement is staffed against — implicit acknowledgement that current processes have gaps significant enough to require dedicated improvement work"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Loan servicing organization",
                "proposed_solution": None,
                "context_tag": "other"
            }
        ]
    },

    # =========================================================
    # JD - SANY CAPITAL USA - 6 rows
    # =========================================================
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 17, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Manual data entry from finance contracts; no auto-extraction",
                "verbatim_quotes": [
                    "Manual data entry from finance contracts. Contract data is not auto-extracted from documents",
                    "analyst re-keys terms into the loan administration system"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan operations analyst",
                "proposed_solution": None,
                "context_tag": "onboarding"
            }
        ]
    },
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 18, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Portfolio reports manually prepared and distributed; no self-serve reporting",
                "verbatim_quotes": [
                    "Portfolio reports are manually prepared and distributed (likely Excel + email), not pulled from a real-time servicing system",
                    "No self-serve reporting layer"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan operations analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 19, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Risk trend identification is manual; no automated early-warning",
                "verbatim_quotes": [
                    "Trend identification is a manual analyst job, not an automated early-warning system on the servicing platform",
                    "Risk signal surfacing is human-dependent"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Loan operations analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 20, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Customers manually UAT vendor upgrades; vendor release quality is poor",
                "verbatim_quotes": [
                    "Analyst performs UAT on vendor upgrades — implies servicing platform releases break things often enough that customers manually validate them",
                    "Quietly devastating signal on vendor release quality"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan operations analyst",
                "proposed_solution": None,
                "context_tag": "support"
            }
        ]
    },
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 21, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Compliance auditing is manual file review; scales with headcount, not loans",
                "verbatim_quotes": [
                    "Compliance auditing is manual file review — no automated compliance rules embedded in the servicing platform",
                    "Compliance scales with headcount, not loans"
                ],
                "severity_signal": "high",
                "dollar_signal": "Compliance scales with headcount",
                "user_segment": "Loan operations analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sany Capital USA Loan Ops", "row": 22, "source_type": "job_description", "persona_fit": 3,
        "pains": [
            # Low-signal data point about platform identity, not pain. Per prompt: ignore non-pain content.
        ]
    },

    # =========================================================
    # JD - SR ASSOCIATE CAPITAL MARKETS - 6 rows
    # =========================================================
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 23, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Master portfolio data tapes manually generated; no API-driven feeds",
                "verbatim_quotes": [
                    "\"Master portfolio data tapes\" = industry shorthand for batch CSV/Excel files exchanged with capital providers",
                    "Implies data tapes are manually generated files, not API-driven feeds",
                    "Persistent data-exchange pain"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Capital markets analyst at lender/servicer",
                "proposed_solution": None,
                "context_tag": "integration"
            }
        ]
    },
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 24, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "KPI reports require eng team + Power BI; no self-serve",
                "verbatim_quotes": [
                    "Producing KPI reports requires internal eng team involvement",
                    "Servicing/portfolio platforms do not expose self-service reporting at the level needed; KPIs require eng + Power BI"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Capital markets analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 25, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "No canonical reconciliation system; org develops procedures from scratch",
                "verbatim_quotes": [
                    "Org is *developing* reconciliation procedures, not buying them off-the-shelf",
                    "No canonical reconciliation system; procedures are home-grown"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Capital markets / finance org",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 26, "source_type": "job_description", "persona_fit": 5,
        "pains": [
            {
                "pain_summary": "Reconciliation drift: capital provider reports don't match business-level reports",
                "verbatim_quotes": [
                    "The fact that this is staffed against implies capital provider reports DO NOT naturally match business-level reports",
                    "Direct reconciliation drift signal"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Capital markets analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 27, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Compliance reporting/monitoring inefficient; manual, fragmented, duplicative",
                "verbatim_quotes": [
                    "\"Handled efficiently\" framed as a goal implies current compliance reporting/monitoring is *not* efficient",
                    "manual, fragmented, or duplicative across capital providers"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Capital markets analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "JD - Sr. Associate Capital Markets", "row": 28, "source_type": "job_description", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Analysts use SQL to work around platform; UI doesn't expose needed data",
                "verbatim_quotes": [
                    "SQL is a hard requirement for an analyst role",
                    "servicing/portfolio platforms do not expose data through native UI in the way the org needs",
                    "Analysts go around the platform via SQL"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Capital markets analyst",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },

    # =========================================================
    # NCINO - Reddit + ForumAggregate
    # =========================================================
    {
        "source_id": "Ncino", "row": 29, "source_type": "Reddit", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Code quality poor; many null pointer exceptions",
                "verbatim_quotes": [
                    "nCino code is also held by duct tape",
                    "so many null pointer exceptions",
                    "nCino is pure trash"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "performance"
            },
            {
                "pain_summary": "Support quality poor; few competent reps",
                "verbatim_quotes": [
                    "their support is not very good",
                    "There's a rare few that are actually helpful/know what they're talking about"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "support"
            },
            {
                "pain_summary": "No trial or technical documentation pre-contract",
                "verbatim_quotes": [
                    "they refuse to provide either a hands-on trial period or any technical documentation at all until AFTER you sign the contract"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "other"
            }
        ]
    },
    {
        "source_id": "Ncino", "row": 30, "source_type": "ForumAggregate", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Underwriter post-migration: nCino is god awful",
                "verbatim_quotes": [
                    "Underwriter here. We just migrated to nCino a year ago and it is god awful"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Underwriter, post-migration",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },

    # =========================================================
    # NORTRIDGE - 8 reviews
    # =========================================================
    {
        "source_id": "Nortridge Loan System", "row": 31, "source_type": "Review", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Awkward implementation of automations and rules",
                "verbatim_quotes": [
                    "Some weird way to accomplish some of the automations or rules"
                ],
                "severity_signal": "low",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 32, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Support is horrible; 6+ months to fix bugs; argues bugs aren't bugs",
                "verbatim_quotes": [
                    "when it comes to true support they cannot keep up",
                    "they either try to argue that it's not a bug or they take 6+ months to fix the bug",
                    "NortRidge's support of their NLS product is horrible"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Licensed user with internal software developers",
                "proposed_solution": None,
                "context_tag": "support"
            },
            {
                "pain_summary": "User Interface is ancient and not easy to use",
                "verbatim_quotes": [
                    "Their User Interface is ancient and is not awful to use, but it's certainly not easy"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "core_workflow"
            },
            {
                "pain_summary": "Shared hosting; one client's setup affects others",
                "verbatim_quotes": [
                    "I would HIGHLY recommend never using NLS provided hosting",
                    "Their hosting product is on a shared environment where one clients setup can affect another clients setup"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Hosted-platform customer",
                "proposed_solution": None,
                "context_tag": "performance"
            },
            {
                "pain_summary": "Vendor moved server and changed DB path without notification",
                "verbatim_quotes": [
                    "they moved it to a new server without telling us and changed the database path without telling us",
                    "Our internal manager asked them multiple times if their update would require and username, password, or server changes and they kept telling him 'No'"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Hosted-platform customer using report server / MSSQL replica",
                "proposed_solution": None,
                "context_tag": "support"
            },
            {
                "pain_summary": "Updates only during business hours; no rollback process",
                "verbatim_quotes": [
                    "They will only do software updates during standard business hours, so they have to take your business offline while they do it",
                    "you have to pray they don't cause new problems after the software updates are done",
                    "We most certainly had issues with bugs or problems after an update and they don't have a rollback process"
                ],
                "severity_signal": "high",
                "dollar_signal": "take your business offline",
                "user_segment": "Hosted-platform customer",
                "proposed_solution": None,
                "context_tag": "support"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 33, "source_type": "Review", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Limited branch security and reporting options",
                "verbatim_quotes": [
                    "The limited options on branch security and reporting"
                ],
                "severity_signal": "low",
                "dollar_signal": None,
                "user_segment": None,
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 34, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Missing GL module",
                "verbatim_quotes": [
                    "Most CFOs would love to have a GL module"
                ],
                "severity_signal": "low",
                "dollar_signal": None,
                "user_segment": "CFO",
                "proposed_solution": "GL module",
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 35, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "App-server latency is biggest staff complaint; no available fix",
                "verbatim_quotes": [
                    "The largest problems we are having as a licensed user of NLS is latency between using the application an it communicating with the server",
                    "The latency is the biggest complaint among staff",
                    "there seems to be no readily available solution to fix this without going higher and higher up the chain of command"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Licensed user; staff/operators",
                "proposed_solution": None,
                "context_tag": "performance"
            },
            {
                "pain_summary": "Customization barrier: undocumented languages, sparse API docs",
                "verbatim_quotes": [
                    "Customization feels like it has a barrier to entry",
                    "The programming languages are not made apparent up front",
                    "the APIs documentation is difficult to find and sometimes lacking in detailed information"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "User trying to customize/integrate",
                "proposed_solution": None,
                "context_tag": "integration"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 36, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Even simple changes required developer support after months of training",
                "verbatim_quotes": [
                    "after months of technical training, even simple changes required developer support",
                    "we paid for many features we never used"
                ],
                "severity_signal": "high",
                "dollar_signal": "we paid for many features we never used",
                "user_segment": "Customer who switched away to Bryt",
                "proposed_solution": "modular pricing / lower technical-support reliance",
                "context_tag": "support"
            }
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 37, "source_type": "Review", "persona_fit": 4,
        "pains": [
            # No clean pain — review is positive, "Cons" section actually contains praise of the product over Point.
        ]
    },
    {
        "source_id": "Nortridge Loan System", "row": 38, "source_type": "Review", "persona_fit": 4,
        "pains": [
            {
                "pain_summary": "Support responsiveness issues; slow feature delivery (e.g., web client)",
                "verbatim_quotes": [
                    "We still have issues with support responsiveness",
                    "Its taken a long time to get a web client but its on the way"
                ],
                "severity_signal": "medium",
                "dollar_signal": None,
                "user_segment": "Loan servicing customer",
                "proposed_solution": None,
                "context_tag": "support"
            },
            {
                "pain_summary": "API integrations limited; improving but not yet adequate",
                "verbatim_quotes": [
                    "We still are waiting for better API integrations but they have come a long way"
                ],
                "severity_signal": "low",
                "dollar_signal": None,
                "user_segment": "Loan servicing customer",
                "proposed_solution": None,
                "context_tag": "integration"
            }
        ]
    },

    # =========================================================
    # TRAZMO - Vendor Perspective (3 rows)
    # =========================================================
    {
        "source_id": "Trazmo", "row": 40, "source_type": "Vendor Perspective", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Reconciliation drift; numbers stop matching; reports lose trust",
                "verbatim_quotes": [
                    "Reconciliation drift, numbers stop matching and no one trusts the reports"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing organizations (vendor articulation)",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Trazmo", "row": 41, "source_type": "Vendor Perspective", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Manual work persists in disbursements, adjustments, restructures, recoveries",
                "verbatim_quotes": [
                    "Manual work creeping back in around disbursements, adjustments, restructures, and recoveries"
                ],
                "severity_signal": "high",
                "dollar_signal": None,
                "user_segment": "Loan servicing organizations (vendor articulation)",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
    {
        "source_id": "Trazmo", "row": 42, "source_type": "Vendor Perspective", "persona_fit": 3,
        "pains": [
            {
                "pain_summary": "Switching trigger: recon becomes full-time job and trust is gone",
                "verbatim_quotes": [
                    "Most teams don't switch systems because they want more features. They switch when recon becomes a full-time job and trust in the numbers is gone"
                ],
                "severity_signal": "high",
                "dollar_signal": "recon becomes a full-time job",
                "user_segment": "Servicing organizations evaluating system switches",
                "proposed_solution": None,
                "context_tag": "core_workflow"
            }
        ]
    },
]

# ============================================================
# OUTPUT 1: Per-source JSON
# ============================================================
out_json = '/sessions/kind-hopeful-edison/mnt/outputs/extraction_results.json'
with open(out_json, 'w') as f:
    json.dump(results, f, indent=2)

# ============================================================
# OUTPUT 2: Flat XLSX — one row per pain
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pains_Flat"

headers = [
    'pain_id', 'source_id', 'source_row', 'source_type', 'persona_fit',
    'pain_summary', 'severity_signal', 'context_tag', 'user_segment',
    'dollar_signal', 'proposed_solution', 'verbatim_quotes'
]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    cell.alignment = Alignment(vertical='center', wrap_text=True)

pain_id = 0
empty_sources = []
for r in results:
    if not r['pains']:
        empty_sources.append((r['source_id'], r['row']))
        continue
    for p in r['pains']:
        pain_id += 1
        ws.append([
            f"P{pain_id:03d}",
            r['source_id'],
            r['row'],
            r['source_type'],
            r['persona_fit'],
            p['pain_summary'],
            p['severity_signal'],
            p['context_tag'],
            p['user_segment'],
            p['dollar_signal'],
            p['proposed_solution'],
            ' || '.join(p['verbatim_quotes'])
        ])

widths = [8, 30, 8, 16, 8, 50, 10, 16, 30, 30, 30, 80]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Add a Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.append(["Metric", "Value"])
ws2.append(["Total source rows", len(results)])
ws2.append(["Total pains extracted", pain_id])
ws2.append(["Sources with 0 pains", len(empty_sources)])
ws2.append(["Pains per source (avg)", round(pain_id / len(results), 2)])
ws2.append([])
ws2.append(["Severity breakdown"])
from collections import Counter
sev = Counter()
ctx = Counter()
src = Counter()
for r in results:
    for p in r['pains']:
        sev[p['severity_signal']] += 1
        ctx[p['context_tag']] += 1
        src[r['source_type']] += 1
for k, v in sev.most_common():
    ws2.append([f"  {k}", v])
ws2.append([])
ws2.append(["Context tag breakdown"])
for k, v in ctx.most_common():
    ws2.append([f"  {k}", v])
ws2.append([])
ws2.append(["Source type breakdown (pain-level)"])
for k, v in src.most_common():
    ws2.append([f"  {k}", v])
ws2.append([])
ws2.append(["Sources with 0 pains (review needed)"])
for sid, srow in empty_sources:
    ws2.append([f"  {sid} (row {srow})"])
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 12

out_xlsx = '/sessions/kind-hopeful-edison/mnt/outputs/extraction_pains_flat.xlsx'
wb.save(out_xlsx)
print(f"Saved JSON: {out_json}")
print(f"Saved XLSX: {out_xlsx}")
print(f"Total pains: {pain_id}")
print(f"Empty sources: {empty_sources}")
print(f"Severity: {dict(sev)}")
print(f"Context: {dict(ctx)}")
